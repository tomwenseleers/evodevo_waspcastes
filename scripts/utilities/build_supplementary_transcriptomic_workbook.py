#!/usr/bin/env python3
"""Build the compact transcriptomic supplementary workbook exported by R."""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEETS = [
    "Table_index",
    "S4_DE_summary",
    "S5A_PLS_axes",
    "S5B_PLS_LOOCV",
    "S6_Fig1_GO",
    "S7_Fig2A_regressions",
    "S8_Fig2B_ridge",
]

INTEGER_RE = re.compile(r"^[+-]?\d+$")
FLOAT_RE = re.compile(
    r"^[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?$"
)


def infer_value(value: str):
    if value == "" or value.upper() == "NA":
        return None
    if INTEGER_RE.fullmatch(value):
        try:
            return int(value)
        except ValueError:
            return value
    if FLOAT_RE.fullmatch(value):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def load_tsv(path: Path) -> list[list[object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter="\t"))
    if not rows:
        raise ValueError(f"Empty table source: {path}")
    return [rows[0]] + [[infer_value(value) for value in row] for row in rows[1:]]


def column_width(values: list[object], header: str) -> float:
    longest = max((len(str(value)) for value in values if value is not None), default=0)
    if header in {"GO_term", "Description"}:
        return min(max(longest + 2, 28), 55)
    if header == "Contributing_HOGs":
        return 55
    if header in {"Species", "Analysis_unit", "Contrast", "Biological_contrast"}:
        return min(max(longest + 2, 16), 30)
    return min(max(longest + 2, 11), 24)


def build(source_dir: Path, output_file: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    navy = "44546A"
    pale_blue = "EAF0F6"
    white = "FFFFFF"
    border_side = Side(style="thin", color="B7C3D0")

    for sheet_index, sheet_name in enumerate(SHEETS, start=1):
        rows = load_tsv(source_dir / f"{sheet_name}.tsv")
        worksheet = workbook.create_sheet(sheet_name)
        for row in rows:
            worksheet.append(row)

        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.row_dimensions[1].height = 32

        for cell in worksheet[1]:
            cell.font = Font(name="Arial", size=10, bold=True, color=white)
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(bottom=border_side)

        for row_number in range(2, worksheet.max_row + 1):
            if row_number % 2 == 0:
                fill = PatternFill("solid", fgColor=pale_blue)
            else:
                fill = PatternFill(fill_type=None)
            for cell in worksheet[row_number]:
                cell.font = Font(name="Arial", size=9)
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    header = worksheet.cell(1, cell.column).value or ""
                    cell.number_format = "0.000E+00" if "P" in str(header) else "0.000"

        headers = [str(cell.value) for cell in worksheet[1]]
        for column_number, header in enumerate(headers, start=1):
            values = [worksheet.cell(row, column_number).value for row in range(1, worksheet.max_row + 1)]
            worksheet.column_dimensions[get_column_letter(column_number)].width = column_width(values, header)

        table = Table(
            displayName=f"SupplementaryTable{sheet_index}",
            ref=worksheet.dimensions,
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)

    # Fail immediately if the package cannot read back its own output.
    checked = load_workbook(output_file, read_only=False, data_only=False)
    if checked.sheetnames != SHEETS:
        raise RuntimeError(f"Unexpected worksheet order: {checked.sheetnames}")
    checked.close()


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: build_supplementary_transcriptomic_workbook.py SOURCE_DIR OUTPUT.xlsx")
    build(Path(sys.argv[1]), Path(sys.argv[2]))


if __name__ == "__main__":
    main()
